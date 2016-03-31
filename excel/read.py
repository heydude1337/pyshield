# -*- coding: utf-8 -*-
"""
Created on Fri Jan  8 12:18:40 2016

@author: 757021
"""

import numpy as np
import openpyxl
import warnings
import datetime

from collections import OrderedDict
EXCEL_EXTENSIONS = ('.xls', '.xlsx', '.xlm', '.xlam')

DATA='Data body'
COLUMN_HEADER = 'Column Header'
ROW_HEADER = 'Row Header'

def read_excel_file(excel_file, make_numpy=True):
  excel_workbook = _open_workbook(excel_file)
  data={}  
  for sheet_name in excel_workbook.get_sheet_names():
    data[sheet_name]=import_excel_sheet(excel_workbook[sheet_name], make_numpy=make_numpy)
  return data

def read_excel_sheet_from_file(excel_file, sheet_name, startrow=1, endrow=-1, startcol=1, endcol=-1, make_numpy=True):
  excel_workbook = _open_workbook(excel_file)
  excel_sheet = excel_workbook[sheet_name]
  data = import_excel_sheet(excel_sheet, startrow=startrow, endrow=endrow, startcol=startcol, endcol=endcol, make_numpy=make_numpy)
  return data

def import_excel_sheet(excel_sheet, startrow=1, endrow=-1, startcol=1, endcol=-1, make_numpy=False):  

  data = read_range(excel_sheet, startrow=startrow, endrow=endrow, startcol=startcol, endcol=endcol)
  if make_numpy: 
    data=_convert_to_numpy(data)    
  return data

def read_range(sheet, startrow=1, endrow=-1, startcol=1, endcol=-1):
  if endrow==-1:
    endrow=sheet.max_row
  if endcol==-1:
    endcol=sheet.max_column

  nrows = endrow-startrow+1
  ncols = endcol-startcol+1
  
  data=[None]*nrows
  
  for i in range(0, nrows):
    rowdata=[None]*ncols
    for j in range(0, ncols):
      rowdata[j]=sheet.cell(row=i+startrow, column=j+startcol).value
    data[i]=rowdata
  
  return data
  
def parse_sheet_data(sheet_data, row_column_header = -1, column_row_header=-1, 
                     make_data_numeric=True, row_name = ROW_HEADER, 
                     column_name = COLUMN_HEADER, data_name = DATA):
  data={}
  row_header = None
  column_header = None
  if row_column_header > -1:
    column_header = sheet_data[row_column_header,column_row_header+1:]        
  if column_row_header > -1:
    row_header = sheet_data[row_column_header+1:, column_row_header]    
  
  data = sheet_data[row_column_header+1:, column_row_header+1:]
  
  if make_data_numeric:
    data=data.astype(float)
      

  if row_column_header > -1 and column_row_header == -1:
    data_dict=OrderedDict() 
    for i, colname in enumerate(column_header):
      data_dict[colname]=data[:,i]     
  elif row_column_header == -1 and column_row_header > -1:
    data_dict=OrderedDict()
    for i, rowname in enumerate(row_header):
      data_dict[rowname]=data[i,:]    
  else:
    data_dict=OrderedDict()
    data_dict[row_name] = row_header
    data_dict[column_name] = column_header
    data_dict[data_name] = data.transpose()
  
  return data_dict

def make_numeric(table, col_names=[], row_names=[], data_type = float):
  if len(col_names)>0:
    names = col_names
  elif len(row_names)>0:
    names=row_names
  
  for name in names:
    if type(table[name][0]) in [datetime.time, datetime.date, datetime.datetime]:
      table[name] = date_time_to_days(table[name], data_type)
    else:    
      table[name]=table[name].astype(data_type)
    
def date_time_to_days(np_values, data_type):
  dtzero = datetime.datetime(1900,1,1,0,0,0)
  dzero = datetime.date(1900,1,1)
  tzero = datetime.time(0,0,0)
  for i, value in enumerate(np_values):
    if type(value)==datetime.time:
      dt=datetime.datetime.combine(dzero, value)-dtzero
    elif type(value)==datetime.date:
      dt=datetime.datetime.combine(value, tzero)-dtzero 
    elif type(value)==datetime.datetime:
      dt=value-dtzero
    np_values[i]=dt.total_seconds()/(24*3600)
  return np_values.astype(data_type)
      
def is_excel(excel_file):
  return excel_file.lower().endswith(EXCEL_EXTENSIONS)
  
def _open_workbook(excel_file):
  return openpyxl.load_workbook(excel_file) 
  
 
def _convert_to_numpy(data):  
  try:
    data=np.array(data)
  except:
    raise
    warnings.warn('Data connot be converted to numpy array')
  return data
    
  
  
  
#  
#    
#    
#  if colheader > 0:
#    for i in range(rowheader+1,table.shape[0]+1):
#      col_header.append(Sheet.cell(column=i, row=colheader).value)
#      
#  if rowheader > 0:  
#    for i in range(colheader+1,table.shape[1]+1):
#      row_header.append(Sheet.cell(column=rowheader, row=i).value)
#    
#  return (col_header, row_header, table)
#
#def remove_zeros(table):
#  while np.all(table[-1,:]==0):
#      table=table[0:-1,:]
#      
#  while np.all(table[:,-1]==0):
#    table=table[:,0:-1]
#  
#  return table
  


