# -*- coding: utf-8 -*-
"""
Created on Wed Jan 27 16:14:49 2016

@author: 757021
"""

import openpyxl


def write_dict_to_excel_sheet(excel_file, sheet_name='sheet1', data_dict={}):
  wb = openpyxl.load_workbook(excel_file)
  if sheet_name not in wb.sheetnames:
    wb.create_sheet(sheet_name)
    
  sheet = wb[sheet_name]

  col_names=data_dict.keys()
  
  for i, col_name in enumerate(col_names):
    sheet.cell(row=1, column=i+1).value = list(col_names)[i]
    
  for i, col_name in enumerate(col_names):
    for j in range(0, len(data_dict[col_name])):
      sheet.cell(column=i+1, row=j+2).value = data_dict[col_name][j]
    
  wb.save(excel_file)
  
  